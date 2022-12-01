from openpyxl import Workbook
wb = Workbook()

ws = wb.active

# ws['A1'] = "Test Data"

testdata = [['Name','City'],['Ira','Sofia'],['Dushyant','Dublin'],['Shyam','Mumbai'],['Sudershan','Noida'],['Rajesh','Bangaluru'],['Sachin','London'],['Aditya','Dublin']]
for data in testdata:
    ws.append(data)
wb.save("Testexcel2.xlsx")


# for i in range(1,6):
#   for j in range(1,5):
#       ws.cell(row=i,column=j).value=i+j


