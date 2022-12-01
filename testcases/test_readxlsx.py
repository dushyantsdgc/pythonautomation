from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='Testexcel.xlsx')
sh = wb['Sheet']

row_ct = sh.max_row
col_ct = sh.max_column

for i in range(1, row_ct+1):
    for j in range(1, col_ct+1):
        print(sh.cell(row=i, column=j).value)








#print(sh['A3'].value)
#print(wb['Demosheet']['A2'].value)
#print(sh.cell(2, 2).value)

